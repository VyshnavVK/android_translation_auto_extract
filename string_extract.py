import xml.etree.ElementTree as ET
from openpyxl import Workbook

def extract_strings(ar_strings_file, en_strings_file, output_xlsx):
    value = ""
    try:

        tree = ET.parse(ar_strings_file)
        root = tree.getroot()
        
        tree2 = ET.parse(en_strings_file)
        root2 = tree2.getroot()


        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Strings"

        ws.cell(row=1, column=1).value = "Name"
        ws.cell(row=1, column=2).value = "English"
        ws.cell(row=1, column=3).value = "Arabic"


        row = 2
        for string in root.findall("string"):
            name = string.attrib.get("name")
            arabic = string.text.strip() if string.text else ""
           
            if string.attrib.get("translatable") == "false":
                value = name
                print(f"skipped : {value}")
                continue

            for en_strings in root2.findall("string"):
               name2 = en_strings.attrib.get("name")
               english = en_strings.text.strip() if en_strings.text else ""
                    
               if name == name2:
                  ws.cell(row=row, column=2).value = english     

            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=3).value = arabic
            row += 1


        wb.save(output_xlsx)

        print(f"Data extracted and saved to '{output_xlsx}'.")

    except FileNotFoundError:
        print(f"Error: '{ar_strings_file}' or {en_strings_file} not found.")
    except ET.ParseError:
        print(f"Error: Failed to parse '{ar_strings_file}' or '{en_strings_file}'. Please check if the XML is well-formed.")
    except Exception as e:
        print(f"An error occurred: {e}")


ar_strings_file = "ar_strings.xml"  
en_strings_file = "en_strings.xml"
output_xlsx = "extracted_strings.xlsx"
extract_strings(ar_strings_file,en_strings_file,output_xlsx)

